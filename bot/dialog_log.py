# -*- coding: utf-8 -*-
"""Журнал диалогов в Postgres + сборка Excel-отчёта (этап 38).

Зачем отдельно от технического лога (этап 26): тот лог — строки в контейнере для
разработчика (request-id, tool-call, тайминги). Здесь — ДОЛГОживущая таблица для
ВЛАДЕЛЬЦА: с кем говорили, что спрашивали, что бот искал, сколько нашёл. Из неё по
кодовому слову собирается .xlsx (три листа: Диалоги / Искали-не-нашли / Сводка).

Три жёстких правила:
- **Запись best-effort.** Любая ошибка БД проглатывается в warning — бот не должен
  падать или тормозить из-за журнала (критерий приёмки этапа 38).
- **Телефонов покупателей тут нет** — мессенджеры их боту не отдают (только имя/ник).
- Пишем ровно то, что нужно отчёту: канал, ключ чата, имя, ник, запрос, что искал
  бот, сколько нашёл, ответ, время.
"""
from __future__ import annotations

import io
from typing import Optional

import asyncpg

from .logger import logger


async def obespechit_tablicu(pool: asyncpg.Pool, schema: str) -> None:
    """Создать таблицу журнала, если её нет (best-effort, вызывать при старте).

    `naydeno = -1` — бот НЕ искал за ход (светская реплика/кнопка); `>= 0` — искал
    и столько нашёл. Так «искали, но не нашли» = `naydeno = 0` (искал и пусто),
    а болтовню (`-1`) в этот раздел не тащим."""
    try:
        async with pool.acquire() as con:
            await con.execute(f"""
                create table if not exists {schema}.dialog_log (
                    id          bigserial primary key,
                    ts          timestamptz not null default now(),
                    channel     text not null,
                    chat_key    text not null,
                    imya        text,
                    nik         text,
                    zapros      text not null,
                    iskal       text,
                    naydeno     integer not null default -1,
                    otvet       text
                );
                create index if not exists dialog_log_ts_idx on {schema}.dialog_log (ts);
            """)
    except Exception as e:  # noqa: BLE001 — журнал не должен ронять старт
        logger.warning(f"Журнал диалогов: не удалось создать таблицу ({e})")


async def zapisat(
    pool: Optional[asyncpg.Pool], schema: str, channel: str, chat_key,
    zapros: str, otvet: str, *, imya: Optional[str] = None,
    nik: Optional[str] = None, iskal: Optional[list[str]] = None,
    naydeno: int = -1,
) -> None:
    """Записать один ход диалога. Best-effort: ошибки — в warning, наружу не летят."""
    if pool is None:
        return
    try:
        iskal_txt = "; ".join(iskal) if iskal else None
        async with pool.acquire() as con:
            await con.execute(
                f"""insert into {schema}.dialog_log
                    (channel, chat_key, imya, nik, zapros, iskal, naydeno, otvet)
                    values ($1,$2,$3,$4,$5,$6,$7,$8)""",
                channel, str(chat_key), imya, nik, zapros, iskal_txt, naydeno, otvet,
            )
    except Exception as e:  # noqa: BLE001
        logger.warning(f"Журнал диалогов: запись не удалась ({e})")


# ─────────────────────────── сборка Excel-отчёта ───────────────────────────

async def postroit_otchet_xlsx(pool: asyncpg.Pool, schema: str) -> bytes:
    """Собрать .xlsx (bytes) из журнала. Три листа: Диалоги / Искали-не-нашли / Сводка.
    Бросает исключение, если БД недоступна — вызывающий ловит и отвечает по-человечески."""
    rows = await pool.fetch(
        f"""select ts, channel, chat_key, imya, nik, zapros, iskal, naydeno, otvet
            from {schema}.dialog_log order by ts"""
    )
    return _v_knigu(rows)


def _v_knigu(rows) -> bytes:
    """Чистая сборка книги из уже прочитанных строк (без БД — удобно тестировать)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def _dt(ts) -> str:
        # ts — timestamptz (aware); формат под человека, без микросекунд
        return ts.strftime("%d.%m.%Y %H:%M") if ts is not None else ""

    def _oформить(ws, zagolovki: list[str], shiriny: list[int]) -> None:
        ws.append(zagolovki)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for i, w in enumerate(shiriny, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ── Лист 1: Диалоги (строка = сообщение) ──
    ws1 = wb.active
    ws1.title = "Диалоги"
    _oформить(ws1,
              ["Дата", "Площадка", "Имя", "Ник", "Запрос клиента",
               "Что искал бот", "Найдено", "Ответ бота"],
              [17, 11, 18, 16, 34, 30, 9, 60])
    for r in rows:
        naydeno = r["naydeno"]
        ws1.append([
            _dt(r["ts"]), r["channel"], r["imya"] or "", r["nik"] or "",
            r["zapros"] or "", r["iskal"] or "",
            "" if naydeno is None or naydeno < 0 else naydeno,
            r["otvet"] or "",
        ])
    for row in ws1.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[7].alignment = Alignment(wrap_text=True, vertical="top")

    # ── Лист 2: Искали, но не нашли (naydeno = 0) ──
    ws2 = wb.create_sheet("Искали, но не нашли")
    _oформить(ws2, ["Дата", "Площадка", "Имя", "Ник", "Запрос", "Что искал бот"],
              [17, 11, 18, 16, 34, 34])
    for r in rows:
        if r["naydeno"] == 0:
            ws2.append([_dt(r["ts"]), r["channel"], r["imya"] or "", r["nik"] or "",
                        r["zapros"] or "", r["iskal"] or ""])

    # ── Лист 3: Сводка ──
    ws3 = wb.create_sheet("Сводка")
    sobesedniki = {(r["channel"], r["chat_key"]) for r in rows}
    vsego = len(rows)
    ne_nashli = sum(1 for r in rows if r["naydeno"] == 0)

    # топ запросов клиента (по нормализованному тексту)
    schet: dict[str, int] = {}
    for r in rows:
        q = (r["zapros"] or "").strip().lower()
        if q:
            schet[q] = schet.get(q, 0) + 1
    top = sorted(schet.items(), key=lambda kv: -kv[1])[:25]

    ws3["A1"] = "Показатель"
    ws3["B1"] = "Значение"
    for c in ws3[1]:
        c.font = Font(bold=True)
    ws3.append(["Всего сообщений", vsego])
    ws3.append(["Уникальных собеседников", len(sobesedniki)])
    ws3.append(["Запросов без результата", ne_nashli])
    ws3.append([])
    ws3.append(["Топ запросов", "Раз"])
    r0 = ws3.max_row
    for c in ws3[r0]:
        c.font = Font(bold=True)
    for q, n in top:
        ws3.append([q, n])
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
